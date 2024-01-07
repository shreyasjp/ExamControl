import csv
import openpyxl
from openpyxl.utils import get_column_letter

class RoomInfo:
    def __init__(self, room_id, room_name, priority, total_seats, filled_seats, rows, columns, matrix, subjects):
        """
        Initialize RoomInfo object with room details.
        """
        self.RoomID = room_id
        self.RoomName = room_name
        self.Priority = priority
        self.TotalSeats = total_seats
        self.FilledSeats = filled_seats
        self.Rows = rows
        self.Columns = columns
        self.Matrix = matrix
        self.Subjects = subjects
        self.IsFilled = self.calculate_is_filled()

    def calculate_is_filled(self):
        """
        Calculate if the room is filled based on filled seats and total seats.
        """
        return self.FilledSeats == self.TotalSeats

import mysql.connector as a
b = a.connect(host='localhost', user='root', passwd='root')
d = b.cursor(buffered=True)
subs = {} #Dictionary to store PRN:Subject pairs
prn = [] #List of all students

# Fetch room details from database
d.execute('SELECT * FROM test.rooms JOIN test.room_class_specs ON test.rooms.RoomClassID = test.room_class_specs.RoomClassID order by test.room_class_specs.Priority desc;')
rooms = [] # List of all rooms and their specifications
for i in d.fetchall():
    matrix = [[0] * 2 * i[7] for _ in range(i[8])]
    rooms.append(RoomInfo(i[0], i[1], i[5], i[6], 0, i[8], i[7], matrix, {}))

# Fetch subject details from database
d.execute('select * from test.prn_sub;')
e = d.column_names
f = list(e) # List of all subjects
f.pop(0) # Remove Sl. No 

for i in f:
    z = 'select `' + i + '` from test.prn_sub where `' + i + '` is not null;'
    d.execute(z)
    rows = d.rowcount
    g = d.fetchall()
    for j in range(len(g)):
        subs[g[j][0]] = i
        prn.append(g[j][0])

totStuds = len(subs) # Total number of students
counter = 0 # Counter to access students

rooms_required = [] # List of all rooms as room objects
seats_available = 0  # Number of seats alloted

# Determine the required rooms based on available seats
for i in range(len(rooms)):
    if totStuds - seats_available <= 10:
        #print('Students Left:', totStuds - seats_available)
        break
    if seats_available < totStuds:
        rooms_required.append(rooms[i])
        seats_available += rooms[i].TotalSeats

temp = [] # Store Room Name, Subjects alloted to room, Start and End PRN

# Assign students to rooms based on priority and available seats
for i in rooms_required:
    if not(i.IsFilled):
        for j in range(2 * i.Columns):
            for k in range(i.Rows):
                if i.FilledSeats <= i.TotalSeats / 2:
                    if j % 2 == 0:
                        room_name = i.RoomName
                        i.Matrix[k][j] = str(prn[counter])

                        if subs[prn[counter]] not in temp or room_name not in temp:
                            temp.append(str(prn[counter-1]))
                            temp.append(room_name)
                            temp.append(subs[prn[counter]])
                            temp.append(str(prn[counter]))
                        counter += 1
                        i.FilledSeats += 1
                else:
                    continue

temp.append(str(prn[counter-1]))
temp = temp[1:]
res = [] # List of Students of each subject alloted to each room

# Generate result data for each assigned student
for i in range(0, len(temp), 4):
    room_name = temp[i]
    subject = temp[i + 1]
    roll_number = temp[i + 2]
    roll_number_ending = temp[i + 3]
    res.append([room_name, f'{subject} = {roll_number} - {roll_number_ending}'])

temp = [] # Store Room Name, Subjects alloted to room, Start and End PRN

# Assign remaining students to rooms
for i in rooms_required:
    if not(i.IsFilled):
        for j in range(2 * i.Columns):
            for k in range(i.Rows):
                if i.Matrix[k][j] == 0:
                    i.Matrix[k][j] = str(prn[counter])
                    room_name = i.RoomName
                    if subs[prn[counter]] not in temp or room_name not in temp:
                        temp.append(str(prn[counter-1]))
                        temp.append(room_name)
                        temp.append(subs[prn[counter]])
                        temp.append(str(prn[counter]))
                    counter += 1
                    i.FilledSeats += 1

temp.append(str(prn[counter]-1))
temp = temp[1:]

# Generate result data for each assigned student
for i in range(0, len(temp), 4):
    room_name = temp[i]
    subject = temp[i + 1]
    roll_number = temp[i + 2]
    roll_number_ending = temp[i + 3]
    res.append([room_name, f'{subject} = {roll_number} - {roll_number_ending}'])

res_sorted = sorted(res, key=lambda x: x[0])

# Create a new Excel workbook to store the data
workbook = openpyxl.Workbook()

# Create the first sheet with data from the first room
first_room = rooms_required[0]
first_sheet = workbook.create_sheet(title=f'0_{first_room.RoomName}')
for row in first_room.Matrix:
    first_sheet.append(row)

# Continue with the rest of the sheets
for index, room in enumerate(rooms_required[1:]):
    sheet = workbook.create_sheet(title=f'{index + 1}_{room.RoomName}')
    for row in room.Matrix:
        sheet.append(row)

# Create two new sheets directly in the existing workbook
room_data_sheet = workbook.create_sheet(title='RoomDataSheet')
students_left_sheet = workbook.create_sheet(title='StudentsLeftSheet')

# Add data to the 'RoomDataSheet' sheet
for data_row in res_sorted:
    room_data_sheet.append(data_row)

# Add data to the 'StudentsLeftSheet' sheet
counter = 0
for i in range(totStuds - seats_available):
    students_left_sheet.append([subs[prn[counter]] + '-' + str(prn[counter])])
    counter += 1

# Remove the default empty sheet
workbook.remove(workbook['Sheet'])

# Set column widths based on the content in each sheet
for sheet in workbook.sheetnames:
    current_sheet = workbook[sheet]
    
    for column in current_sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        current_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Save the modified workbook
excel_filename = "Splitting/RoomAllocationDetails.xlsx"
workbook.save(excel_filename)

print(f'Excel file updated successfully: {excel_filename}')

# CSV codes

"""
for room in rooms_required:
    filename = f'{rooms_required.index(room)}_{room.RoomName}_matrix.csv'
    filename= "Splitting/"+filename
    
    with open(filename, 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile, delimiter='\t')

        for row in range(room.Rows):
            csv_writer.writerow(room.Matrix[row])

    print(f'CSV file created for {room.RoomName}: {filename}') 
"""

""" 
with open('room_data.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    for i in res_sorted:
        writer.writerow(i) 
"""

""" 
for i in range(totStuds-seats_available):
    with open('StudentsLeft.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow([subs[prn[counter]]+'-'+str(prn[counter])])
    counter+=1 
"""